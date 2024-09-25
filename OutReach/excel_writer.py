import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
import win32com.client
import time
import pythoncom 

#sets a generic variable to None for isinstance tests        
NoneType = type(None)

class Excel_Writer():
    def __init__(self, excel_location, impacted_sheets):
        self.excel_location = excel_location
        self.impacted_sheets = impacted_sheets
        # self.one_drive_url = r"https://trinitysolarsys-my.sharepoint.com/personal/joshuabeach_trinity-solar_com/_layouts/15/onedrive.aspx?id=%2Fpersonal%2Fjoshuabeach%5Ftrinity%2Dsolar%5Fcom%2FDocuments&view=0"
        # self.sheet_name = "Outreach Coverage Tracker.xlsx"
        

    def refresh_sheet(self):
        """
        Refreshes formula data in excel sheet
        Interactively waits for the data to finish updating before closing
        """
        xlapp = win32com.client.DispatchEx("Excel.Application", pythoncom.CoInitialize())
        xlapp.DisplayAlerts = True
        wb = xlapp.Workbooks.Open(self.excel_location)
        wb.RefreshAll()
        xlapp.CalculateUntilAsyncQueriesDone()
        wb.Save()
        wb.Close(SaveChanges=True)
        xlapp.Quit()
        pythoncom.CoUninitialize()

        return True
    
    def unprotect_workbook(self) -> bool:
        """
        Unprotects each sheet in the excel workbook
        **run after download..
        """
        ExcelWorkbook = openpyxl.load_workbook(filename=self.excel_location)
        for sheet in ExcelWorkbook.sheetnames:
            if sheet in self.impacted_sheets:
                print(f"Disabling Protection on {sheet}")

                ws = ExcelWorkbook.get_sheet_by_name(sheet)
                ws.protection.disable()

        ExcelWorkbook.save(filename=self.excel_location)
        ExcelWorkbook.close()
        
        return True
    
    def protect_workbook(self) -> bool:
        """
        Protects the Excel Workbook using set password (chooses sheets)
        """
        ExcelWorkbook = openpyxl.load_workbook(self.excel_location)
        for sheet in ExcelWorkbook.sheetnames:
            if sheet in self.impacted_sheets:
                ws = ExcelWorkbook.get_sheet_by_name(sheet)
                print(f"Enabling Protection on {sheet}")

                ws.protection.sheet = True
                ws.protection.set_password('Trinity2')
                # ws.protection.set_password('trinity1')
            
        ExcelWorkbook.save(self.excel_location)
        ExcelWorkbook.close()

        return True
    
    def load_excel_workbook(self, sheet) -> pd.DataFrame:
        """
        Loads the Excel sheet for use
        """
        wb = load_workbook(filename=self.excel_location)
        sheet_ranges = wb[sheet]
        df = pd.DataFrame(sheet_ranges.values)

        return df

    def Variables(self, df=None):
        """
        Styles the variables sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Variables')
        sheet['A11'] = datetime.now()
        file.save(self.excel_location)

        return True
    
    def WT(self, df):
        """
        Writes to WT sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('WT')

        sheet_counter = 2
        for counter, row in enumerate(df['Lead Generator Office']):
            
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Lead Generator'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Created Date'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Demo Date'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Demo Completed'].iloc[counter]
            sheet[f'F{sheet_counter}'] = df['Zip/Postal Code'].iloc[counter]

            sheet_counter += 1
        while len(str(sheet[f'C{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            sheet[f'F{sheet_counter}'] = ""
            
            sheet_counter += 1

        file.save(self.excel_location)
        return True
    
    def Demo(self, df):
        """
        Writes to Demo sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Demo')

        sheet_counter = 2
        for counter, row in enumerate(df['Opportunity: Lead Generator Office']):

            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Opportunity: Lead Generator: Full Name'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Demo Date'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Zip'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Demo Completed'].iloc[counter]
            sheet[f'F{sheet_counter}'] = df['Direct Lead ID'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'C{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            sheet[f'F{sheet_counter}'] = ""

            sheet_counter += 1

        file.save(self.excel_location)
        
        return True
    
    def Demo_30_Days(self, df):
        """
        Writes to Demo Last 30 days
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Demo 30 Days')

        sheet_counter = 2
        
        for counter, row in enumerate(df['Opportunity: Lead Generator Office']):

            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Opportunity: Lead Generator: Full Name'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Demo Date'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Zip'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Demo Completed'].iloc[counter]
            sheet[f'F{sheet_counter}'] = df['Direct Lead ID'].iloc[counter]
            sheet[f'G{sheet_counter}'] = df['Appointment Completed'].iloc[counter]


            sheet_counter += 1

        while len(str(sheet[f'C{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            sheet[f'F{sheet_counter}'] = ""
            sheet[f'G{sheet_counter}'] = ""
            
            sheet_counter += 1

        file.save(self.excel_location)

        return True 
    
    def Contracts(self, df):
        """
        Writes to Contracts sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Contracts')

        sheet_counter = 2
        for counter, row in enumerate(df['Lead Generator Office']):
            
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Lead Generator: Full Name'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Contract Signed'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Account Name: Billing Zip/Postal Code'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'C{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            
            sheet_counter += 1

        file.save(self.excel_location)
        
        return True
    
    def No_Rep(self, df):
        """
        Writes to No Rep sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('No Rep')

        sheet_counter = 2
        for counter, row in enumerate(df['Direct Lead: Opportunity']):
            
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Job: Job Name'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Start'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Outcome'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Appointment Outcome'].iloc[counter]
            sheet[f'F{sheet_counter}'] = df['Outcome Reason'].iloc[counter]
            sheet[f'G{sheet_counter}'] = df['Customer Confirmation Status'].iloc[counter]
            sheet[f'H{sheet_counter}'] = df['Region'].iloc[counter]
            sheet[f'I{sheet_counter}'] = df['Direct Lead: Zip'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'C{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            #print(sheet_counter)
            #print(f"Erasing extra value... Row: {sheet_counter}, value={sheet[f'A{sheet_counter}'].value}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            sheet[f'F{sheet_counter}'] = ""
            sheet[f'G{sheet_counter}'] = ""
            sheet[f'H{sheet_counter}'] = ""
            sheet[f'I{sheet_counter}'] = ""

            sheet_counter += 1

        file.save(self.excel_location)
        return True
    
    def SameDay(self, df):
        """
        Writes to the Same Day sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('SameDay')
        
        sheet_counter = 2
        for counter, row in enumerate(df['State']):
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Customer Confirmation Status'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Created Date'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Start'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Status'].iloc[counter]
            sheet[f'F{sheet_counter}'] = df['Job Status'].iloc[counter]
            sheet[f'G{sheet_counter}'] = df['Outcome'].iloc[counter]
            sheet[f'H{sheet_counter}'] = df['Outcome Reason'].iloc[counter]
            sheet[f'I{sheet_counter}'] = df['Created via Button'].iloc[counter]
            sheet[f'J{sheet_counter}'] = df['Zip'].iloc[counter]
            sheet[f'K{sheet_counter}'] = df['Opportunity: Lead Generator Office'].iloc[counter]
            sheet[f'L{sheet_counter}'] = df['Opportunity: Lead Generator: Full Name'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'D{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            sheet[f'F{sheet_counter}'] = ""
            sheet[f'G{sheet_counter}'] = ""
            sheet[f'H{sheet_counter}'] = ""
            sheet[f'I{sheet_counter}'] = ""
            sheet[f'J{sheet_counter}'] = ""
            sheet[f'K{sheet_counter}'] = ""
            sheet[f'L{sheet_counter}'] = ""
            
            sheet_counter += 1

        file.save(self.excel_location)

        return True

    def Contact(self, df):
        """
        Writes to the contact sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Contact')

        sheet_counter = 2
        for counter, row in enumerate(df['Name Formula']):
            
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Division'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Sales Office'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Start Date'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'D{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            
            sheet_counter += 1
            
        file.save(self.excel_location)

        return True
    
    def User(self, df):
        """
        Writes to the User sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('User')
        print(df)
        sheet_counter = 2
        for counter, row in enumerate(df['Full Name']):
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Division'].iloc[counter]
            sheet[f'C{sheet_counter}'] = df['Sales Office'].iloc[counter]
            sheet[f'D{sheet_counter}'] = df['Start Date'].iloc[counter]
            sheet[f'E{sheet_counter}'] = df['Full User ID'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'D{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            sheet[f'C{sheet_counter}'] = ""
            sheet[f'D{sheet_counter}'] = ""
            sheet[f'E{sheet_counter}'] = ""
            
            sheet_counter += 1
            
        file.save(self.excel_location)

        return True

    def Office_Sort(self, df):
        """
        Writes to Office Sort sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Office Sort')

        sheet_counter = 13
        for counter, row in enumerate(df['Outreach Sales Offices']):
            
            sheet[f'B{sheet_counter}'] = row

            sheet_counter += 1

        while len(str(sheet[f'B{sheet_counter}'].value)) >= 1:
            if not sheet[f'C{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'B{sheet_counter}'] = ""
    
            sheet_counter += 1

        file.save(self.excel_location)

        return True

    def Skedulo_Data(self, df):
        """
        Writes to the Skedulo Data sheet
        """
        file = openpyxl.load_workbook(self.excel_location)
        sheet = file.get_sheet_by_name('Skedulo Data')

        sheet_counter = 2
        for counter, row in enumerate(df['Resource Name']):
            
            sheet[f'A{sheet_counter}'] = row
            sheet[f'B{sheet_counter}'] = df['Tag: Tag'].iloc[counter]

            sheet_counter += 1

        while len(str(sheet[f'A{sheet_counter}'].value)) >= 1:
            if not sheet[f'A{sheet_counter}'].value:
                break
            print(f"Erasing extra value... Row: {sheet_counter}")
            sheet[f'A{sheet_counter}'] = ""
            sheet[f'B{sheet_counter}'] = ""
            
            sheet_counter += 1
            
        file.save(self.excel_location)

        return True
